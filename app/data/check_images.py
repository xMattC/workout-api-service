from pathlib import Path
import re

from openpyxl import load_workbook

"""
Validate exercise image coverage.

Checks that each exercise in the spreadsheet has corresponding image files
in the local image directory and reports missing entries.
"""


def slugify(value):
    """Convert exercise name into image filename slug.

    Normalises text into lowercase hyphenated format used for image naming.
    """
    text = str(value).strip().lower()
    text = text.replace("_", "-")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def image_exists(images_dir, slug, index):
    """Check if an image exists for a given slug and index.

    Supports multiple common image extensions.
    """
    extensions = [".jpg", ".jpeg", ".png", ".webp"]
    base = f"{slug}-{index}"

    for ext in extensions:
        if (images_dir / f"{base}{ext}").exists():
            return True

    return False


def find_column(sheet, name):
    """Locate a column index by header name.

    Performs a case-insensitive match on the first row of the sheet.
    """
    for cell in sheet[1]:
        if str(cell.value).strip().lower() == name.lower():
            return cell.column

    raise ValueError(f"Column '{name}' not found")


def check_missing_images(spreadsheet_path, images_dir):
    """Check spreadsheet exercises for missing image files.

    Iterates through each exercise, validates image presence, and prints a
    summary of missing image entries.
    """
    wb = load_workbook(spreadsheet_path)
    ws = wb.active

    name_col = find_column(ws, "name")

    missing = []
    missing_both = []

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value

        if not name:
            continue

        slug = slugify(name)
        print(f"Checking: {name} (slug: {slug})")

        has_1 = image_exists(images_dir, slug, 1)
        has_2 = image_exists(images_dir, slug, 2)

        if not has_1 or not has_2:
            if not has_1 and not has_2:
                missing_both.append(slug)
                missing.append(slug)
            elif not has_1:
                missing.append(f"{slug}-1")
            elif not has_2:
                missing.append(f"{slug}-2")

    print("\n--- Missing Images ---")
    for item in sorted(missing):
        print(item)

    print("\n--- Summary ---")
    print(f"Total missing entries: {len(missing)}")
    print(f"Exercises missing both images: {len(missing_both)}")


def main():
    """Run image validation against local dataset.

    Uses the local spreadsheet and images directory to verify dataset completeness.
    """
    base_dir = Path(__file__).resolve().parent

    spreadsheet = base_dir / "full_exercise_sheet.xlsx"
    images_dir = base_dir / "images"

    check_missing_images(spreadsheet, images_dir)


if __name__ == "__main__":
    main()
