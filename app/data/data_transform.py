import json
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

"""Prepare exercise dataset assets from free-exercise-db.

Builds a spreadsheet for manual curation and consolidates images into a
single directory with consistent naming.

Source:
    https://github.com/yuhonas/free-exercise-db/
"""

SHEET_HEADERS = [
    "keep",
    "name",
    "difficulty",
    "equipment",
    "category",
    "primary_muscles",
    "secondary_muscles",
    "instructions",
]

COLUMN_WIDTHS = [10, 35, 14, 18, 18, 28, 35, 120]


def join_list(values, separator=" | "):
    """Join list values into a spreadsheet-friendly string.

    Used to flatten list fields (muscles, instructions) into a single cell.
    """
    if not values:
        return ""

    cleaned_values = [str(value).strip() for value in values if str(value).strip()]
    return separator.join(cleaned_values)


def build_row(exercise):
    """Convert a single exercise record into a spreadsheet row.

    Extracts relevant fields from the source JSON and formats them for Excel.
    """
    return [
        "yes",
        exercise.get("name", "").strip(),
        exercise.get("level", "").strip(),
        exercise.get("equipment", "").strip(),
        exercise.get("category", "").strip(),
        join_list(exercise.get("primaryMuscles", [])),
        join_list(exercise.get("secondaryMuscles", [])),
        join_list(exercise.get("instructions", []), separator=" || "),
    ]


def format_sheet(worksheet):
    """Apply formatting to the Excel worksheet.

    Styles header row, sets column widths, and freezes the header.
    """
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for column_index, width in enumerate(COLUMN_WIDTHS, start=1):
        column_letter = chr(64 + column_index)
        worksheet.column_dimensions[column_letter].width = width

    worksheet.freeze_panes = "A2"


def build_from_directory(input_dir, output_file):
    """Build a spreadsheet from raw exercise JSON files.

    Iterates through all JSON files in the source directory and writes a
    flattened dataset to an Excel file for manual review.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Exercises"
    worksheet.append(SHEET_HEADERS)

    valid_count = 0
    skipped_count = 0

    for json_file in sorted(Path(input_dir).rglob("*.json")):
        try:
            with json_file.open("r", encoding="utf-8") as file:
                data = json.load(file)

            if not isinstance(data, dict) or "name" not in data:
                skipped_count += 1
                continue

            worksheet.append(build_row(data))
            valid_count += 1

        except Exception as exc:
            skipped_count += 1
            print(f"Skipping invalid JSON file: {json_file} ({exc})")

    format_sheet(worksheet)
    workbook.save(output_file)

    print(f"Spreadsheet created: {output_file}")
    print(f"Valid exercises written: {valid_count}")
    print(f"Skipped files: {skipped_count}")


def slugify_folder_name(name):
    """Convert a folder name into a slug for image filenames.

    Used to standardise image naming across all exercises.
    """
    cleaned = name.strip().lower()
    cleaned = cleaned.replace("_", "-")
    cleaned = re.sub(r"[^a-z0-9-]+", "-", cleaned)
    cleaned = re.sub(r"-+", "-", cleaned).strip("-")
    return cleaned


def copy_image(source_file, destination_file):
    """Copy an image file if it exists.

    Ensures destination directory exists before copying.
    """
    if not source_file.exists():
        return False

    destination_file.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_file, destination_file)
    return True


def collect_images(input_dir, output_dir):
    """Collect and rename exercise images into a single directory.

    Copies 0.jpg and 1.jpg from each exercise folder and renames them
    using a slug-based convention.
    """
    input_path = Path(input_dir)
    output_path = Path(output_dir)

    output_path.mkdir(parents=True, exist_ok=True)

    copied_count = 0
    missing_count = 0

    for exercise_dir in sorted(input_path.iterdir()):
        if not exercise_dir.is_dir():
            continue

        slug = slugify_folder_name(exercise_dir.name)

        source_1 = exercise_dir / "0.jpg"
        source_2 = exercise_dir / "1.jpg"

        dest_1 = output_path / f"{slug}-1.jpg"
        dest_2 = output_path / f"{slug}-2.jpg"

        if copy_image(source_1, dest_1):
            copied_count += 1
        else:
            missing_count += 1
            print(f"Missing image: {source_1}")

        if copy_image(source_2, dest_2):
            copied_count += 1
        else:
            missing_count += 1
            print(f"Missing image: {source_2}")

    print(f"Image collection complete: {output_path}")
    print(f"Images copied: {copied_count}")
    print(f"Missing images: {missing_count}")


def main():
    """Run the data preparation pipeline.

    Builds the exercise spreadsheet and consolidates images into a
    single directory for later seeding into the application.
    """
    base_dir = Path(__file__).resolve().parent

    source_exercises_dir = base_dir / "exercises"
    output_sheet_path = base_dir / "exercise_sheet.xlsx"
    output_images_dir = base_dir / "images"

    if not source_exercises_dir.exists():
        raise FileNotFoundError(f"Source directory not found: {source_exercises_dir}")

    build_from_directory(input_dir=source_exercises_dir, output_file=output_sheet_path)
    collect_images(input_dir=source_exercises_dir, output_dir=output_images_dir)


if __name__ == "__main__":
    main()
